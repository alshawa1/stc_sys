import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

def generate_styled_daily_excel(port_table, sup_df, col_df, df_pay_filtered, report_date, contact_table=None):
    """
    Generates a compact, executive one-screen Dashboard Excel Report.
    Features:
    - Tight, fitted table column widths matching text exactly without excessive gaps.
    - Clear, non-overlapping charts:
      * Horizontal Bar Chart for Collection Amounts (displays exact amounts cleanly at the end of each bar).
      * Vertical Column Chart for Collection Rate % (outEnd percentages above columns).
      * Executive Pie Chart for Contact Status Breakdown (percentages & categories).
      * Horizontal Bar Chart for Top 5 Collectors.
    """
    wb = Workbook()
    
    # ── Color Palette ──
    HEADER_FILL     = PatternFill(start_color="0F2027", end_color="0F2027", fill_type="solid")
    ACCENT_FILL     = PatternFill(start_color="005F4B", end_color="005F4B", fill_type="solid")
    CONTACT_FILL    = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    KPI_CARD_FILL   = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
    TOTAL_FILL      = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    HEADER_FONT     = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    TITLE_FONT      = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    SECTION_FONT    = Font(name="Segoe UI", size=12, bold=True, color="005F4B")
    TOTAL_FONT      = Font(name="Segoe UI", size=10, bold=True, color="004D40")
    BOLD_FONT       = Font(name="Segoe UI", size=10, bold=True)

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    KPI_BORDER = Border(
        left=Side(style='medium', color='005F4B'),
        right=Side(style='medium', color='005F4B'),
        top=Side(style='medium', color='005F4B'),
        bottom=Side(style='medium', color='005F4B')
    )

    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_LEFT   = Alignment(horizontal='left', vertical='center')
    ALIGN_RIGHT  = Alignment(horizontal='right', vertical='center')

    # ════════════════════════════════════════════════════════════════
    # MAIN SHEET: 📊 الداشبورد التنفيذية (Executive Single Page)
    # ════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "الداشبورد التنفيذية"
    ws.views.sheetView[0].rightToLeft = True

    # ── 1. Top Header Banner ──
    ws.merge_cells("A1:Q1")
    banner_cell = ws["A1"]
    banner_cell.value = f"📈 التقرير اليومي التنفيذي — فولو اب ({report_date}) | نظام مهاره للتحصيل"
    banner_cell.fill = PatternFill(start_color="004D40", end_color="004D40", fill_type="solid")
    banner_cell.font = TITLE_FONT
    banner_cell.alignment = ALIGN_CENTER

    # ── 2. Top KPI Cards Row ──
    tot_debt = port_table[port_table['المحفظة'] == '📊 الإجمالي']['إجمالي المديونية'].values[0] if '📊 الإجمالي' in port_table['المحفظة'].values and 'إجمالي المديونية' in port_table.columns else port_table['إجمالي المديونية'].sum()
    tot_coll = port_table[port_table['المحفظة'] == '📊 الإجمالي']['إجمالي التحصيل'].values[0] if '📊 الإجمالي' in port_table['المحفظة'].values and 'إجمالي التحصيل' in port_table.columns else port_table['إجمالي التحصيل'].sum()
    tot_cust = port_table[port_table['المحفظة'] == '📊 الإجمالي']['عدد العملاء'].values[0] if '📊 الإجمالي' in port_table['المحفظة'].values and 'عدد العملاء' in port_table.columns else port_table['عدد العملاء'].sum()
    coll_pct = (tot_coll / tot_debt * 100) if tot_debt > 0 else 0

    cnt_contacted     = contact_table[contact_table['المحفظة'] == '📊 الإجمالي']['تم التوصل'].values[0] if contact_table is not None and not contact_table.empty and 'تم التوصل' in contact_table.columns else 0
    cnt_no_ans_closed = contact_table[contact_table['المحفظة'] == '📊 الإجمالي']['لا يرد ومغلق'].values[0] if contact_table is not None and not contact_table.empty and 'لا يرد ومغلق' in contact_table.columns else 0
    cnt_other         = contact_table[contact_table['المحفظة'] == '📊 الإجمالي']['عدم توصل - أخرى'].values[0] if contact_table is not None and not contact_table.empty and 'عدم توصل - أخرى' in contact_table.columns else 0
    cnt_rate          = contact_table[contact_table['المحفظة'] == '📊 الإجمالي']['نسبة تم التوصل %'].values[0] if contact_table is not None and not contact_table.empty and 'نسبة تم التوصل %' in contact_table.columns else 0

    kpis = [
        ("👥 العملاء", f"{tot_cust:,.0f}", "A3:B4"),
        ("💰 المديونية", f"{tot_debt:,.0f} ﷼", "C3:D4"),
        ("💵 التحصيل", f"{tot_coll:,.0f} ﷼", "E3:F4"),
        ("📈 التحصيل %", f"{coll_pct:.1f}%", "G3:H4"),
        ("📞 تم التوصل", f"{cnt_contacted:,.0f}", "I3:K4"),
        ("📵 لا يرد ومغلق", f"{cnt_no_ans_closed:,.0f}", "L3:N4"),
        ("📈 التوصل %", f"{cnt_rate:.1f}%", "O3:P4"),
    ]

    for title, val_str, range_str in kpis:
        ws.merge_cells(range_str)
        top_left_col = range_str.split(":")[0]
        c = ws[top_left_col]
        c.value = f"{title}\n{val_str}"
        c.fill = KPI_CARD_FILL
        c.font = Font(name="Segoe UI", size=10, bold=True, color="004D40")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for row_cells in ws[range_str]:
            for cell_item in row_cells:
                cell_item.border = KPI_BORDER

    # ── 3. Section Title: Portfolio Summary Table ──
    ws.cell(row=6, column=1, value="📋 ملخص أداء المحافظ والتحصيل").font = SECTION_FONT

    headers1 = list(port_table.columns)
    for c_idx, h in enumerate(headers1, 1):
        cell = ws.cell(row=8, column=c_idx, value=h)
        cell.fill = ACCENT_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    start_row = 9
    num_rows = len(port_table)
    
    for r_idx, row in port_table.iterrows():
        row_num = start_row + r_idx
        is_total_row = (row_num == start_row + num_rows - 1) or ("إجمالي" in str(row.iloc[0]))
        
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=c_idx)
            c_name = headers1[c_idx - 1]
            
            if isinstance(val, (int, float, np.integer, np.floating)):
                if "%" in c_name:
                    cell.value = float(val) / 100.0 if val > 1 else float(val)
                    cell.number_format = '0.0%'
                elif "مديونية" in c_name or "تحصيل" in c_name:
                    cell.value = float(val)
                    cell.number_format = '#,##0 "﷼"'
                else:
                    cell.value = int(val) if val == int(val) else float(val)
                    cell.number_format = '#,##0'
            else:
                cell.value = str(val)

            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER if c_idx > 1 else ALIGN_RIGHT

            if is_total_row:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

    table1_end_row = start_row + num_rows - 1

    # ── Collection Bar Chart 1 & 2 (Horizontal Bar Chart for Collection, Column Chart for Rate %) ──
    try:
        data_len = num_rows - 1 if num_rows > 1 else 1
        data_ref = Reference(ws, min_col=4, min_row=8, max_row=8 + data_len)
        cats_ref = Reference(ws, min_col=1, min_row=9, max_row=8 + data_len)

        # Horizontal Bar Chart for Collection Amount (Guarantees zero text overlap for large amounts)
        chart1 = BarChart()
        chart1.type = "bar"
        chart1.style = 10
        chart1.title = "💰 إجمالي التحصيل حسب المحفظة"
        chart1.width = 15
        chart1.height = 8.5
        chart1.legend = None
        
        chart1.dataLabels = DataLabelList()
        chart1.dataLabels.showVal = True
        chart1.dataLabels.position = "outEnd"

        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        ws.add_chart(chart1, "I8")

        if "نسبة التحصيل %" in headers1:
            rate_col_idx = headers1.index("نسبة التحصيل %") + 1
            chart2 = BarChart()
            chart2.type = "col"
            chart2.style = 11
            chart2.title = "📈 نسبة التحصيل %"
            chart2.width = 14
            chart2.height = 8.5
            chart2.legend = None

            chart2.dataLabels = DataLabelList()
            chart2.dataLabels.showVal = True
            chart2.dataLabels.position = "outEnd"

            data_ref2 = Reference(ws, min_col=rate_col_idx, min_row=8, max_row=8 + data_len)
            chart2.add_data(data_ref2, titles_from_data=True)
            chart2.set_categories(cats_ref)
            ws.add_chart(chart2, "Q8")
    except Exception:
        pass

    # ── 4. Section 2: Contact Breakdown Table & Executive Pie Chart ──
    r_sec2 = max(table1_end_row + 2, 18)
    ws.cell(row=r_sec2, column=1, value="📞 تحليل حالات التوصل وعدم التوصل (تم التوصل - لا يرد ومغلق)").font = SECTION_FONT

    r_cnt_start = r_sec2 + 2
    cnt_end_row = r_cnt_start

    if contact_table is not None and not contact_table.empty:
        cnt_headers = list(contact_table.columns)
        for c_idx, h in enumerate(cnt_headers, 1):
            cell = ws.cell(row=r_cnt_start, column=c_idx, value=h)
            cell.fill = CONTACT_FILL
            cell.font = HEADER_FONT
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

        for r_idx, row in contact_table.iterrows():
            row_num = r_cnt_start + 1 + r_idx
            is_tot = "إجمالي" in str(row.iloc[0])
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=c_idx)
                c_n = cnt_headers[c_idx - 1]
                if isinstance(val, (int, float, np.integer, np.floating)):
                    if "%" in c_n:
                        cell.value = float(val) / 100.0 if val > 1 else float(val)
                        cell.number_format = '0.0%'
                    else:
                        cell.value = int(val) if val == int(val) else float(val)
                        cell.number_format = '#,##0'
                else:
                    cell.value = str(val)

                cell.border = THIN_BORDER
                cell.alignment = ALIGN_CENTER if c_idx > 1 else ALIGN_RIGHT

                if is_tot:
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="1E3A8A")

        cnt_num_rows = len(contact_table)
        cnt_end_row = r_cnt_start + cnt_num_rows

        # Offscreen Helper Data for Contact Pie Chart
        ws['Y1'] = 'الحالة'
        ws['Z1'] = 'العدد'
        ws['Y2'] = 'تم التوصل'
        ws['Z2'] = float(cnt_contacted)
        ws['Y3'] = 'لا يرد ومغلق'
        ws['Z3'] = float(cnt_no_ans_closed)
        ws['Y4'] = 'عدم توصل - أخرى'
        ws['Z4'] = float(cnt_other)

        # Executive Pie Chart for Contact Breakdown
        try:
            pie_chart = PieChart()
            pie_chart.title = "📞 توزيع حالات التوصل الإجمالية"
            pie_chart.width = 14
            pie_chart.height = 8.5

            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True

            data_pie = Reference(ws, min_col=26, min_row=1, max_row=4) # Col Z
            cats_pie = Reference(ws, min_col=25, min_row=2, max_row=4) # Col Y
            pie_chart.add_data(data_pie, titles_from_data=True)
            pie_chart.set_categories(cats_pie)

            ws.add_chart(pie_chart, f"I{r_cnt_start}")
        except Exception:
            pass

    # ── 5. Section 3: Top 5 Collectors & Top Supervisors Side-by-Side ──
    r_sec3 = max(cnt_end_row + 2, 28)
    ws.cell(row=r_sec3, column=1, value="🏆 أفضل 5 محصلين وأفضل المشرفين أداءً").font = SECTION_FONT

    r_tables = r_sec3 + 2

    # Left: ⭐ TOP 5 COLLECTORS TABLE
    col_headers = ["الترتيب", "المحصل", "إجمالي التحصيل", "المعدل %"]
    for c_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=r_tables, column=c_idx, value=h)
        cell.fill = ACCENT_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    r_col_curr = r_tables + 1
    top5_cols = col_df.head(5) if col_df is not None and not col_df.empty else pd.DataFrame()
    medals = {1: "🥇 الأول", 2: "🥈 الثاني", 3: "🥉 الثالث", 4: "🏅 الرابع", 5: "🏅 الخامس"}

    if not top5_cols.empty:
        for r_i, r_data in top5_cols.reset_index(drop=True).iterrows():
            rank_num = r_i + 1
            ws.cell(row=r_col_curr, column=1, value=medals.get(rank_num, f"#{rank_num}")).alignment = ALIGN_CENTER
            ws.cell(row=r_col_curr, column=2, value=str(r_data.get('المحصل', ''))).alignment = ALIGN_RIGHT

            c_val = ws.cell(row=r_col_curr, column=3, value=float(r_data.get('إجمالي التحصيل', 0)))
            c_val.number_format = '#,##0 "﷼"'
            c_val.alignment = ALIGN_CENTER

            p_val = ws.cell(row=r_col_curr, column=4, value=float(r_data.get('المعدل %', 0)) / 100.0)
            p_val.number_format = '0.0%'
            p_val.alignment = ALIGN_CENTER

            if rank_num == 1:
                row_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif rank_num == 2:
                row_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            elif rank_num == 3:
                row_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
            else:
                row_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

            for c in range(1, 5):
                cell_item = ws.cell(row=r_col_curr, column=c)
                cell_item.fill = row_fill
                cell_item.border = THIN_BORDER
                cell_item.font = BOLD_FONT
            r_col_curr += 1

    # Right: 👤 TOP SUPERVISORS TABLE
    sup_headers = ["#", "المشرف", "إجمالي التحصيل", "المعدل %"]
    for c_idx, h in enumerate(sup_headers, 6):
        cell = ws.cell(row=r_tables, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    r_sup_curr = r_tables + 1
    if sup_df is not None and not sup_df.empty:
        for r_i, r_data in sup_df.head(5).iterrows():
            ws.cell(row=r_sup_curr, column=6, value=r_data.get('#', r_i+1)).alignment = ALIGN_CENTER
            ws.cell(row=r_sup_curr, column=7, value=str(r_data.get('المشرف', ''))).alignment = ALIGN_RIGHT

            c_val = ws.cell(row=r_sup_curr, column=8, value=float(r_data.get('إجمالي التحصيل', 0)))
            c_val.number_format = '#,##0 "﷼"'
            c_val.alignment = ALIGN_CENTER

            p_val = ws.cell(row=r_sup_curr, column=9, value=float(r_data.get('المعدل %', 0)) / 100.0)
            p_val.number_format = '0.0%'
            p_val.alignment = ALIGN_CENTER

            for c in range(6, 10):
                ws.cell(row=r_sup_curr, column=c).border = THIN_BORDER
            r_sup_curr += 1

    # Chart 4: Top 5 Collectors Horizontal Bar Chart with DATA LABELS
    try:
        if not top5_cols.empty:
            chart3 = BarChart()
            chart3.type = "bar"
            chart3.style = 13
            chart3.title = "⭐ أداء أفضل 5 محصلين"
            chart3.width = 13
            chart3.height = 6.5
            chart3.legend = None

            chart3.dataLabels = DataLabelList()
            chart3.dataLabels.showVal = True
            chart3.dataLabels.position = "outEnd"

            data_ref3 = Reference(ws, min_col=3, min_row=r_tables, max_row=r_tables+len(top5_cols))
            cats_ref3 = Reference(ws, min_col=2, min_row=r_tables+1, max_row=r_tables+len(top5_cols))
            chart3.add_data(data_ref3, titles_from_data=True)
            chart3.set_categories(cats_ref3)
            ws.add_chart(chart3, "K" + str(r_tables))
    except Exception:
        pass

    # Column Auto-Width Adjustment (Tight & Compact matching text exactly)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column > 20:
            continue # Don't format offscreen helper columns Y and Z
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 9), 18)

    # ════════════════════════════════════════════════════════════════
    # SHEET 2: 💳 السدادات التفصيلية
    # ════════════════════════════════════════════════════════════════
    ws_pay = wb.create_sheet(title="السدادات التفصيلية")
    ws_pay.views.sheetView[0].rightToLeft = True

    df_export_pay = df_pay_filtered.drop(columns=['_pay_date', 'm_amt'], errors='ignore')
    pay_headers = list(df_export_pay.columns)

    for c_idx, h in enumerate(pay_headers, 1):
        cell = ws_pay.cell(row=1, column=c_idx, value=h)
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER

    for r_idx, row in df_export_pay.iterrows():
        row_num = r_idx + 2
        for c_idx, val in enumerate(row, 1):
            cell = ws_pay.cell(row=row_num, column=c_idx, value=str(val) if pd.notna(val) else '')
            cell.alignment = ALIGN_CENTER

    for col in ws_pay.columns:
        col_letter = get_column_letter(col[0].column)
        ws_pay.column_dimensions[col_letter].width = 15

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
